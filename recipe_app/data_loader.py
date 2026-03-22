from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .highlighting import resolve_sentence_descriptors
from .models import (
    DescriptorMatch,
    IngredientLine,
    RecipeChunk,
    RecipeDocument,
    RecipeSentence,
    RecipeStep,
)
from .text_utils import (
    looks_like_recipe_id,
    safe_text,
    to_bool_flag,
    to_int,
    to_optional_text,
)


@dataclass(slots=True)
class RecipeStore:
    workbook_path: Path
    recipes: dict[str, RecipeDocument]

    def get(self, recipe_id: str) -> RecipeDocument:
        return self.recipes[recipe_id]

    def list_recipes(self) -> list[RecipeDocument]:
        return sorted(self.recipes.values(), key=lambda recipe: (recipe.title.casefold(), recipe.recipe_id))

    def search(self, query: str) -> list[RecipeDocument]:
        normalized_query = safe_text(query).casefold()
        recipes = self.list_recipes()
        if not normalized_query:
            return recipes
        return [recipe for recipe in recipes if normalized_query in recipe.title.casefold()]

    def filter_recipes(
        self,
        *,
        title_query: str = "",
        recipe_id_query: str = "",
        categories: list[str] | None = None,
        rating_range: tuple[float, float] | None = None,
        descriptor_codes: list[str] | None = None,
    ) -> list[RecipeDocument]:
        recipes = self.list_recipes()
        normalized_title = safe_text(title_query).casefold()
        normalized_recipe_id = safe_text(recipe_id_query).casefold()
        allowed_categories = set(categories or [])
        required_descriptor_codes = set(descriptor_codes or [])

        filtered: list[RecipeDocument] = []
        for recipe in recipes:
            if normalized_title and normalized_title not in recipe.title.casefold():
                continue
            if normalized_recipe_id and normalized_recipe_id not in recipe.recipe_id.casefold():
                continue
            if categories is not None and recipe.category not in allowed_categories:
                continue
            if rating_range is not None:
                minimum_rating, maximum_rating = rating_range
                if recipe.star_rating is None or not (minimum_rating <= recipe.star_rating <= maximum_rating):
                    continue
            if required_descriptor_codes and not (required_descriptor_codes & set(recipe.descriptor_code_counts)):
                continue
            filtered.append(recipe)
        return filtered


def load_recipe_store(workbook_path: str | Path) -> RecipeStore:
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, data_only=True)

    metadata_by_id = _load_metadata(workbook["Recipe_Metadata"])
    ingredients_by_id, version_two_ingredients_by_id = _load_ingredients(workbook["Ingredients_List"])
    sentences_by_id, version_two_sentences_by_id = _load_sentences(workbook["Procedural_Text"])
    descriptors_by_key = _load_descriptors(workbook["Descriptor_Coding"])

    recipes: dict[str, RecipeDocument] = {}
    for recipe_id, metadata in metadata_by_id.items():
        ingredient_lines = ingredients_by_id.get(recipe_id, [])
        version_two_ingredient_lines = version_two_ingredients_by_id.get(recipe_id, ingredient_lines)
        step_map = defaultdict(list)
        version_two_step_map = defaultdict(list)
        all_descriptors: list[DescriptorMatch] = []

        for sentence in sentences_by_id.get(recipe_id, []):
            descriptor_key = (recipe_id, sentence.step_number, sentence.sentence_number)
            resolved_descriptors = resolve_sentence_descriptors(
                sentence.text,
                descriptors_by_key.get(descriptor_key, []),
            )
            step_map[sentence.step_number].append(
                RecipeSentence(
                    step_number=sentence.step_number,
                    sentence_number=sentence.sentence_number,
                    text=sentence.text,
                    descriptors=resolved_descriptors,
                )
            )
            all_descriptors.extend(resolved_descriptors)

        version_two_source_sentences = version_two_sentences_by_id.get(recipe_id)
        if version_two_source_sentences is None:
            version_two_steps = []
        else:
            for sentence in version_two_source_sentences:
                version_two_step_map[sentence.step_number].append(
                    RecipeSentence(
                        step_number=sentence.step_number,
                        sentence_number=sentence.sentence_number,
                        text=sentence.text,
                        descriptors=[],
                    )
                )

        steps = [
            RecipeStep(
                step_number=step_number,
                sentences=sorted(
                    sentences,
                    key=lambda sentence: sentence.sentence_number,
                ),
            )
            for step_number, sentences in sorted(step_map.items(), key=lambda item: item[0])
        ]
        version_two_steps = (
            [
                RecipeStep(
                    step_number=step_number,
                    sentences=sorted(
                        sentences,
                        key=lambda sentence: sentence.sentence_number,
                    ),
                )
                for step_number, sentences in sorted(version_two_step_map.items(), key=lambda item: item[0])
            ]
            if version_two_step_map
            else steps
        )

        chatbot_context = _build_chatbot_context(metadata, ingredient_lines, steps, all_descriptors)
        chunks = _build_chunks(metadata, ingredient_lines, steps)
        descriptor_code_counts = Counter(descriptor.category_code for descriptor in all_descriptors)
        recipes[recipe_id] = RecipeDocument(
            recipe_id=recipe_id,
            title=metadata["title"],
            category=metadata["category"],
            url=metadata["url"],
            star_rating=metadata["star_rating"],
            review_count=metadata["review_count"],
            ingredients=ingredient_lines,
            steps=steps,
            descriptors=all_descriptors,
            descriptor_count=len(all_descriptors),
            descriptor_code_counts=dict(descriptor_code_counts),
            chatbot_context=chatbot_context,
            chunks=chunks,
            version_two_ingredients=version_two_ingredient_lines,
            version_two_steps=version_two_steps,
        )

    return RecipeStore(workbook_path=workbook_path, recipes=recipes)


def build_recipe_document(
    *,
    recipe_id: str,
    title: str,
    category: str,
    url: str = "",
    star_rating: float | None = None,
    review_count: int | None = None,
    ingredient_lines: list[str] | None = None,
    step_lines: list[str] | None = None,
    version_two_ingredient_lines: list[str] | None = None,
    version_two_step_lines: list[str] | None = None,
) -> RecipeDocument:
    metadata = {
        "recipe_id": recipe_id,
        "title": title,
        "category": category,
        "url": url,
        "star_rating": star_rating,
        "review_count": review_count,
    }
    ingredients = [
        IngredientLine(line_number=index, full_text=line)
        for index, line in enumerate(ingredient_lines or [], start=1)
    ]
    version_two_ingredients = [
        IngredientLine(line_number=index, full_text=line)
        for index, line in enumerate(version_two_ingredient_lines or ingredient_lines or [], start=1)
    ]
    steps = [
        RecipeStep(
            step_number=index,
            sentences=[
                RecipeSentence(
                    step_number=index,
                    sentence_number=1,
                    text=line,
                    descriptors=[],
                )
            ],
        )
        for index, line in enumerate(step_lines or [], start=1)
    ]
    version_two_steps = [
        RecipeStep(
            step_number=index,
            sentences=[
                RecipeSentence(
                    step_number=index,
                    sentence_number=1,
                    text=line,
                    descriptors=[],
                )
            ],
        )
        for index, line in enumerate(version_two_step_lines or step_lines or [], start=1)
    ]
    descriptors: list[DescriptorMatch] = []
    chatbot_context = _build_chatbot_context(metadata, ingredients, steps, descriptors)
    chunks = _build_chunks(metadata, ingredients, steps)
    return RecipeDocument(
        recipe_id=recipe_id,
        title=title,
        category=category,
        url=url,
        star_rating=star_rating,
        review_count=review_count,
        ingredients=ingredients,
        steps=steps,
        descriptors=descriptors,
        descriptor_count=0,
        descriptor_code_counts={},
        chatbot_context=chatbot_context,
        chunks=chunks,
        version_two_ingredients=version_two_ingredients or ingredients,
        version_two_steps=version_two_steps or steps,
    )


def _load_metadata(worksheet) -> dict[str, dict[str, object]]:
    metadata: dict[str, dict[str, object]] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        recipe_id = safe_text(row[0])
        if not looks_like_recipe_id(recipe_id):
            continue
        metadata[recipe_id] = {
            "recipe_id": recipe_id,
            "category": safe_text(row[1]),
            "title": safe_text(row[2]),
            "url": safe_text(row[3]),
            "star_rating": _to_optional_float(row[4]),
            "review_count": _to_optional_int(row[5]),
        }
    return metadata


def _load_ingredients(worksheet) -> tuple[dict[str, list[IngredientLine]], dict[str, list[IngredientLine]]]:
    header_map = _header_index_map(worksheet)
    version_two_text_index = _find_header_index(
        header_map,
        (
            "version_2_ingredient_text",
            "full_ingredient_text_version_2",
            "full_ingredient_text_v2",
            "ingredient_text_version_2",
        ),
    )
    ingredients_by_id: dict[str, list[IngredientLine]] = defaultdict(list)
    version_two_ingredients_by_id: dict[str, list[IngredientLine]] = defaultdict(list)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        recipe_id = safe_text(row[0])
        if not looks_like_recipe_id(recipe_id):
            continue
        ingredient = IngredientLine(
            line_number=to_int(row[1]),
            full_text=safe_text(row[2]),
            quantity=to_optional_text(row[3]),
            unit=to_optional_text(row[4]),
            ingredient_name=to_optional_text(row[5]),
            notes=to_optional_text(row[6]),
        )
        ingredients_by_id[recipe_id].append(ingredient)
        version_two_full_text = (
            safe_text(row[version_two_text_index])
            if version_two_text_index is not None and version_two_text_index < len(row)
            else ""
        )
        if version_two_full_text:
            version_two_ingredients_by_id[recipe_id].append(
                IngredientLine(
                    line_number=ingredient.line_number,
                    full_text=version_two_full_text,
                    quantity=ingredient.quantity,
                    unit=ingredient.unit,
                    ingredient_name=ingredient.ingredient_name,
                    notes=ingredient.notes,
                )
            )
    for lines in ingredients_by_id.values():
        lines.sort(key=lambda item: item.line_number)
    for lines in version_two_ingredients_by_id.values():
        lines.sort(key=lambda item: item.line_number)
    return ingredients_by_id, version_two_ingredients_by_id


def _load_sentences(worksheet) -> tuple[dict[str, list[RecipeSentence]], dict[str, list[RecipeSentence]]]:
    header_map = _header_index_map(worksheet)
    version_two_sentence_index = _find_header_index(
        header_map,
        (
            "version_2_sentence_text",
            "full_sentence_text_version_2",
            "full_sentence_text_v2",
            "sentence_text_version_2",
        ),
    )
    sentences_by_id: dict[str, list[RecipeSentence]] = defaultdict(list)
    version_two_sentences_by_id: dict[str, list[RecipeSentence]] = defaultdict(list)
    seen_sentence_keys: set[tuple[str, int, int, str]] = set()
    seen_version_two_sentence_keys: set[tuple[str, int, int, str]] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        recipe_id = safe_text(row[0])
        if not looks_like_recipe_id(recipe_id):
            continue
        step_number = to_int(row[1])
        sentence_number = to_int(row[2])
        sentence_text = safe_text(row[3])
        if not sentence_text:
            continue
        sentence_key = (recipe_id, step_number, sentence_number, sentence_text)
        if sentence_key in seen_sentence_keys:
            continue
        seen_sentence_keys.add(sentence_key)
        sentences_by_id[recipe_id].append(
            RecipeSentence(
                step_number=step_number,
                sentence_number=sentence_number,
                text=sentence_text,
            )
        )
        version_two_sentence_text = (
            safe_text(row[version_two_sentence_index])
            if version_two_sentence_index is not None and version_two_sentence_index < len(row)
            else ""
        )
        if version_two_sentence_text:
            version_two_sentence_key = (recipe_id, step_number, sentence_number, version_two_sentence_text)
            if version_two_sentence_key not in seen_version_two_sentence_keys:
                seen_version_two_sentence_keys.add(version_two_sentence_key)
                version_two_sentences_by_id[recipe_id].append(
                    RecipeSentence(
                        step_number=step_number,
                        sentence_number=sentence_number,
                        text=version_two_sentence_text,
                    )
                )
    for sentences in sentences_by_id.values():
        sentences.sort(key=lambda sentence: (sentence.step_number, sentence.sentence_number))
    for sentences in version_two_sentences_by_id.values():
        sentences.sort(key=lambda sentence: (sentence.step_number, sentence.sentence_number))
    return sentences_by_id, version_two_sentences_by_id


def _load_descriptors(worksheet) -> dict[tuple[str, int, int], list[DescriptorMatch]]:
    descriptors_by_key: dict[tuple[str, int, int], list[DescriptorMatch]] = defaultdict(list)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        recipe_id = safe_text(row[0])
        if not looks_like_recipe_id(recipe_id):
            continue
        descriptor_text = safe_text(row[3])
        if not descriptor_text:
            continue
        key = (recipe_id, to_int(row[1]), to_int(row[2]))
        descriptors_by_key[key].append(
            DescriptorMatch(
                step_number=key[1],
                sentence_number=key[2],
                descriptor_text=descriptor_text,
                category_code=safe_text(row[4]) or "OTHER_VDD",
                multimodal_flag=to_bool_flag(row[5]),
                redundant_flag=to_bool_flag(row[6]),
            )
        )
    return descriptors_by_key


def _build_chatbot_context(
    metadata: dict[str, object],
    ingredients: list[IngredientLine],
    steps: list[RecipeStep],
    descriptors: list[DescriptorMatch],
) -> str:
    ingredients_block = "\n".join(
        f"- {ingredient.full_text}"
        for ingredient in ingredients
    )
    steps_block = "\n".join(
        f"Step {step.step_number}: "
        + " ".join(sentence.text for sentence in step.sentences)
        for step in steps
    )
    descriptor_block = "\n".join(
        f"- Step {descriptor.step_number}, sentence {descriptor.sentence_number}: "
        f"{descriptor.descriptor_text} [{descriptor.category_code}]"
        for descriptor in descriptors
    )
    return (
        f"Recipe ID: {metadata['recipe_id']}\n"
        f"Title: {metadata['title']}\n"
        f"Category: {metadata['category']}\n"
        f"Star rating: {metadata['star_rating'] or 'Unknown'}\n"
        f"Review count: {metadata['review_count'] or 'Unknown'}\n"
        f"Source URL: {metadata['url']}\n\n"
        "Ingredients:\n"
        f"{ingredients_block or '- No ingredients recorded.'}\n\n"
        "Steps:\n"
        f"{steps_block or 'No steps recorded.'}\n\n"
        "Visual descriptor annotations:\n"
        f"{descriptor_block or '- None recorded.'}"
    )


def _build_chunks(
    metadata: dict[str, object],
    ingredients: list[IngredientLine],
    steps: list[RecipeStep],
) -> list[RecipeChunk]:
    chunks = [
        RecipeChunk(
            chunk_id=f"{metadata['recipe_id']}:overview",
            title="Recipe overview",
            text=(
                f"Title: {metadata['title']}\n"
                f"Category: {metadata['category']}\n"
                f"Star rating: {metadata['star_rating'] or 'Unknown'}\n"
                f"Source URL: {metadata['url']}"
            ),
        ),
        RecipeChunk(
            chunk_id=f"{metadata['recipe_id']}:ingredients",
            title="Ingredients",
            text="\n".join(f"- {ingredient.full_text}" for ingredient in ingredients)
            or "- No ingredients recorded.",
        ),
    ]

    for step in steps:
        step_text = " ".join(sentence.text for sentence in step.sentences)
        descriptor_notes = [
            descriptor
            for sentence in step.sentences
            for descriptor in sentence.descriptors
        ]
        descriptor_lines = "\n".join(
            f"- {descriptor.descriptor_text} [{descriptor.category_code}]"
            for descriptor in descriptor_notes
        )
        chunks.append(
            RecipeChunk(
                chunk_id=f"{metadata['recipe_id']}:step:{step.step_number}",
                title=f"Step {step.step_number}",
                text=(
                    f"Step {step.step_number}\n"
                    f"Instruction: {step_text}\n"
                    "Visual-descriptor annotations:\n"
                    f"{descriptor_lines or '- None recorded.'}"
                ),
            )
        )
    return chunks


def _to_optional_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def _to_optional_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    return int(float(str(value).strip()))


def _header_index_map(worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, value in enumerate(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))):
        normalized = _normalize_header(value)
        if normalized:
            header_map[normalized] = index
    return header_map


def _find_header_index(header_map: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for candidate in candidates:
        if candidate in header_map:
            return header_map[candidate]
    return None


def _normalize_header(value: object) -> str:
    text = safe_text(value).casefold()
    return "".join(character if character.isalnum() else "_" for character in text).strip("_")
